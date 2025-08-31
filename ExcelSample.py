import openpyxl

book = openpyxl.load_workbook("D:\\Projects\\Python_Selenium\\Python_demo.xlsx")
sheet = book.active
cell = sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=10,column=1).value = "Azhar"
book.save("D:\\Projects\\Python_Selenium\\Python_demo.xlsx")

for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i,column=j).value)